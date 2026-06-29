"""
Create ongc_schema_v2 in PostgreSQL and seed from Excel files.
This schema is NOT connected to any website — just for mentor review in pgAdmin.

Usage: python3 seed_v2_schema.py
"""

import openpyxl
from sqlalchemy import create_engine, schema, Column, Integer, String, Float, DateTime, func, MetaData, Table
from sqlalchemy.orm import sessionmaker, DeclarativeBase

DATABASE_URL = "postgresql+psycopg2://ongc_user:ongc_pass@localhost:5433/ongc_db"
SCHEMA = "ongc_schema_v2"

engine = create_engine(DATABASE_URL, echo=False)

# Create schema
with engine.connect() as conn:
    conn.execute(schema.CreateSchema(SCHEMA, if_not_exists=True))
    conn.commit()

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


class Base(DeclarativeBase):
    pass


class AcquisitionTarget(Base):
    __tablename__ = "acquisition_targets"
    __table_args__ = {"schema": SCHEMA}

    id = Column(Integer, primary_key=True, index=True)
    project_name = Column(String, nullable=False, index=True)
    financial_year = Column(String, nullable=False)
    type = Column(String, nullable=False)
    apr = Column(Float, default=0)
    may = Column(Float, default=0)
    jun = Column(Float, default=0)
    jul = Column(Float, default=0)
    aug = Column(Float, default=0)
    sep = Column(Float, default=0)
    oct = Column(Float, default=0)
    nov = Column(Float, default=0)
    dec = Column(Float, default=0)
    jan = Column(Float, default=0)
    feb = Column(Float, default=0)
    mar = Column(Float, default=0)
    total = Column(Float, default=0)
    created_at = Column(DateTime, server_default=func.now())


class Manpower(Base):
    __tablename__ = "manpower"
    __table_args__ = {"schema": SCHEMA}

    id = Column(Integer, primary_key=True, index=True)
    section = Column(String, nullable=False, index=True)
    sl_no = Column(Integer, nullable=True)
    cpf_no = Column(String, nullable=True)
    name = Column(String, nullable=True)
    designation = Column(String, nullable=True)
    mobile = Column(String, nullable=True)
    level = Column(String, nullable=True)
    crc = Column(String, nullable=True)
    assignment = Column(String, nullable=True)
    created_at = Column(DateTime, server_default=func.now())


Base.metadata.create_all(bind=engine)

# ── Seed Acquisition Targets ──

SECTION_NAMES = {"GP-03", "GP-06", "GP-15", "GP-16", "GP-26", "GP-36", "GP-61", "GP-81", "RCC", "REL", "Base Office"}

ACQUISITION_FILE = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/2026.04.20_Monthly Acquisition Targets for  RE2026-27_BE2027-28_032908.xlsx"
MANPOWER_FILE = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/Manpower_Geophysical_Services-2026.06.22_035425.xlsx"

db = SessionLocal()

# Check if already seeded
if db.query(AcquisitionTarget).count() == 0:
    wb = openpyxl.load_workbook(ACQUISITION_FILE, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = None
        sheet_year = None
        for part in sheet_name.split():
            if part in ("BE", "RE"):
                sheet_type = part
            elif "-" in part and len(part) == 7:
                sheet_year = part
        if not sheet_type or not sheet_year:
            print(f"  Skipping sheet '{sheet_name}'")
            continue

        print(f"  Processing: {sheet_name}")
        for row in ws.iter_rows(min_row=3, values_only=True):
            sl, project = row[0], row[1]
            if sl is None or project is None or not isinstance(sl, (int, float)):
                continue
            monthly = [float(v) if v is not None else 0 for v in row[3:15]]
            annual = float(row[2]) if row[2] is not None else 0
            total = float(row[15]) if row[15] is not None else 0
            db.add(AcquisitionTarget(
                project_name=str(project).strip(),
                financial_year=sheet_year,
                type=sheet_type,
                apr=monthly[0], may=monthly[1], jun=monthly[2],
                jul=monthly[3], aug=monthly[4], sep=monthly[5],
                oct=monthly[6], nov=monthly[7], dec=monthly[8],
                jan=monthly[9], feb=monthly[10], mar=monthly[11],
                total=total or annual,
            ))
    db.commit()
    print(f"  Seeded {db.query(AcquisitionTarget).count()} acquisition targets")
else:
    print(f"  Acquisition targets already seeded ({db.query(AcquisitionTarget).count()} records)")

# ── Seed Manpower ──

if db.query(Manpower).count() == 0:
    wb = openpyxl.load_workbook(MANPOWER_FILE, data_only=True)
    ws = wb["Manpower Status GPS-2026"]
    current_section = None
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        a = str(row[0]).strip() if row[0] is not None else ""
        if a in SECTION_NAMES:
            current_section = a
            continue
        if a in ("Sl. No.", "", "Sl"):
            continue
        try:
            sl_no = int(float(a))
        except (ValueError, TypeError):
            continue
        name = str(row[2]).strip() if row[2] is not None else ""
        if not name:
            continue
        db.add(Manpower(
            section=current_section or "Unknown",
            sl_no=sl_no,
            cpf_no=str(row[1]).strip() if row[1] else "",
            name=name,
            designation=str(row[3]).strip() if row[3] else "",
            mobile=str(row[4]).strip() if row[4] else "",
            level=str(row[5]).strip() if row[5] else "",
            crc=str(row[6]).strip() if row[6] else "",
            assignment=str(row[7]).strip() if row[7] else "",
        ))
        count += 1
    db.commit()
    print(f"  Seeded {count} manpower records")
else:
    print(f"  Manpower already seeded ({db.query(Manpower).count()} records)")

db.close()
print("Done! Check pgAdmin → ongc_db → Schemas → ongc_schema_v2 → Tables")
