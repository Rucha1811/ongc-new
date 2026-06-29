"""
Seed the budget database from the two Excel files.
Usage: python seed_data.py
"""

import sys
sys.path.insert(0, ".")

import openpyxl
from app.database import SessionLocal, engine, Base
from app.models import AcquisitionTarget, Manpower

Base.metadata.create_all(bind=engine)

SECTION_NAMES = {"GP-03", "GP-06", "GP-15", "GP-16", "GP-26", "GP-36", "GP-61", "GP-81", "RCC", "REL", "Base Office"}

MONTH_COLS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]

ACQUISITION_FILE = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/2026.04.20_Monthly Acquisition Targets for  RE2026-27_BE2027-28_032908.xlsx"
MANPOWER_FILE = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/Manpower_Geophysical_Services-2026.06.22_035425.xlsx"


def seed_acquisition_targets():
    wb = openpyxl.load_workbook(ACQUISITION_FILE, data_only=True)
    db = SessionLocal()
    try:
        existing = db.query(AcquisitionTarget).count()
        if existing:
            print(f"Skipping acquisition targets — {existing} records already exist.")
            return

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Parse type and financial_year from sheet name
            sheet_type = None
            sheet_year = None
            for part in sheet_name.split():
                if part in ("BE", "RE"):
                    sheet_type = part
                elif "-" in part and len(part) == 7:
                    sheet_year = part
            if not sheet_type or not sheet_year:
                print(f"  Skipping sheet '{sheet_name}' — could not parse type/year")
                continue

            print(f"  Processing sheet: {sheet_name} (type={sheet_type}, year={sheet_year})")

            rows = list(ws.iter_rows(min_row=3, values_only=True))
            for row in rows:
                sl = row[0]
                project = row[1]
                annual = row[2]
                monthly = row[3:15]
                total = row[15]

                if sl is None or project is None:
                    continue
                if not isinstance(sl, (int, float)):
                    continue
                if not str(project).strip():
                    continue

                monthly_vals = []
                for i, m in enumerate(monthly):
                    try:
                        monthly_vals.append(float(m) if m is not None else 0)
                    except (ValueError, TypeError):
                        monthly_vals.append(0)

                try:
                    annual_val = float(annual) if annual is not None else 0
                except (ValueError, TypeError):
                    annual_val = 0
                try:
                    total_val = float(total) if total is not None else 0
                except (ValueError, TypeError):
                    total_val = 0

                record = AcquisitionTarget(
                    project_name=str(project).strip(),
                    financial_year=sheet_year,
                    type=sheet_type,
                    apr=monthly_vals[0],
                    may=monthly_vals[1],
                    jun=monthly_vals[2],
                    jul=monthly_vals[3],
                    aug=monthly_vals[4],
                    sep=monthly_vals[5],
                    oct=monthly_vals[6],
                    nov=monthly_vals[7],
                    dec=monthly_vals[8],
                    jan=monthly_vals[9],
                    feb=monthly_vals[10],
                    mar=monthly_vals[11],
                    total=total_val or annual_val,
                )
                db.add(record)

        db.commit()
        count = db.query(AcquisitionTarget).count()
        print(f"  Seeded {count} acquisition targets")
    finally:
        db.close()


def seed_manpower():
    wb = openpyxl.load_workbook(MANPOWER_FILE, data_only=True)
    ws = wb["Manpower Status GPS-2026"]
    db = SessionLocal()
    try:
        existing = db.query(Manpower).count()
        if existing:
            print(f"Skipping manpower — {existing} records already exist.")
            return

        current_section = None
        count = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            a = str(row[0]).strip() if row[0] is not None else ""

            # Section header
            if a in SECTION_NAMES:
                current_section = a
                continue

            # Skip header rows and empty
            if a in ("Sl. No.", "", "Sl"):
                continue

            # Try to parse as employee number
            try:
                sl_no = int(float(a))
            except (ValueError, TypeError):
                continue

            cpf = str(row[1]).strip() if row[1] is not None else ""
            name = str(row[2]).strip() if row[2] is not None else ""
            designation = str(row[3]).strip() if row[3] is not None else ""
            mobile = str(row[4]).strip() if row[4] is not None else ""
            level = str(row[5]).strip() if row[5] is not None else ""
            crc = str(row[6]).strip() if row[6] is not None else ""
            assignment = str(row[7]).strip() if row[7] is not None else ""

            if not name:
                continue

            record = Manpower(
                section=current_section or "Unknown",
                sl_no=sl_no,
                cpf_no=cpf,
                name=name,
                designation=designation,
                mobile=mobile,
                level=level,
                crc=crc,
                assignment=assignment,
            )
            db.add(record)
            count += 1

        db.commit()
        print(f"  Seeded {count} manpower records")
    finally:
        db.close()


if __name__ == "__main__":
    print("Seeding acquisition targets...")
    seed_acquisition_targets()
    print("Seeding manpower...")
    seed_manpower()
    print("Done!")
