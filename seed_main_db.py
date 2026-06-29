"""
Seed acquisition_targets and manpower_employees into the main database.
Run from the ongc-portal/backend directory.

Usage:
  cd ongc-portal/backend
  python ../../seed_main_db.py
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "ongc-portal", "backend"))
os.chdir(os.path.join(os.path.dirname(__file__), "ongc-portal", "backend"))

import asyncio
import openpyxl
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import text
from app.models.base import AcquisitionTarget, ManpowerEmployee, Base

DATABASE_URL = f"postgresql+asyncpg://ongc_user:ongc_pass@localhost:5433/ongc_db"

engine = create_async_engine(DATABASE_URL, echo=False)
SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

SECTION_NAMES = {"GP-03", "GP-06", "GP-15", "GP-16", "GP-26", "GP-36", "GP-61", "GP-81", "RCC", "REL", "Base Office"}
ACQUISITION_FILE = os.path.join(os.path.dirname(__file__), "2026.04.20_Monthly Acquisition Targets for  RE2026-27_BE2027-28_032908.xlsx")
MANPOWER_FILE = os.path.join(os.path.dirname(__file__), "Manpower_Geophysical_Services-2026.06.22_035425.xlsx")


async def seed():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    db = SessionLocal()

    try:
        # ── Acquisition Targets ──
        existing = await db.execute(text("SELECT count(*) FROM acquisition_targets"))
        if existing.scalar() == 0:
            wb = openpyxl.load_workbook(ACQUISITION_FILE, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_type = None
                sheet_year = None
                for part in sheet_name.split():
                    if part in ("BE", "RE"):
                        sheet_type = part
                    elif "-" in part and len(part) == 7:
                        sheet_year = part
                if not sheet_type or not sheet_year:
                    continue
                print(f"  Processing: {sheet_name}")
                for row in ws.iter_rows(min_row=3, values_only=True):
                    sl, project = row[0], row[1]
                    if sl is None or project is None or not isinstance(sl, (int, float)):
                        continue
                    monthly = [float(v) if v is not None else 0 for v in row[3:15]]
                    annual = float(row[2]) if row[2] is not None else 0
                    total = float(row[15]) if row[15] is not None else 0
                    db.add(AcquisitionTarget(
                        project_name=str(project).strip(),
                        financial_year=sheet_year,
                        type=sheet_type,
                        basin="WON Basin",
                        apr=monthly[0], may=monthly[1], jun=monthly[2],
                        jul=monthly[3], aug=monthly[4], sep=monthly[5],
                        oct=monthly[6], nov=monthly[7], dec=monthly[8],
                        jan=monthly[9], feb=monthly[10], mar=monthly[11],
                        total=total or annual,
                    ))
            await db.commit()
            cnt = await db.execute(text("SELECT count(*) FROM acquisition_targets"))
            print(f"  Seeded {cnt.scalar()} acquisition targets")
        else:
            print("  Acquisition targets already seeded")

        # ── Manpower Employees ──
        existing = await db.execute(text("SELECT count(*) FROM manpower_employees"))
        if existing.scalar() == 0:
            wb = openpyxl.load_workbook(MANPOWER_FILE, data_only=True)
            ws = wb["Manpower Status GPS-2026"]
            current_section = None
            count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                a = str(row[0]).strip() if row[0] is not None else ""
                if a in SECTION_NAMES:
                    current_section = a
                    continue
                if a in ("Sl. No.", "", "Sl"):
                    continue
                try:
                    sl_no = int(float(a))
                except (ValueError, TypeError):
                    continue
                name = str(row[2]).strip() if row[2] is not None else ""
                if not name:
                    continue
                db.add(ManpowerEmployee(
                    section=current_section or "Unknown",
                    basin="WON Basin",
                    sl_no=sl_no,
                    cpf_no=str(row[1]).strip() if row[1] else "",
                    name=name,
                    designation=str(row[3]).strip() if row[3] else "",
                    mobile=str(row[4]).strip() if row[4] else "",
                    level=str(row[5]).strip() if row[5] else "",
                    crc=str(row[6]).strip() if row[6] else "",
                    assignment=str(row[7]).strip() if row[7] else "",
                ))
                count += 1
            await db.commit()
            print(f"  Seeded {count} manpower employees")
        else:
            print("  Manpower employees already seeded")

    finally:
        await db.close()
    await engine.dispose()
    print("Done!")


asyncio.run(seed())
