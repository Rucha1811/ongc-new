import asyncio, re, os, sys
from datetime import date
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from app.models.base import Project, Lookup
import openpyxl

EXCEL_PATH = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/GPS_VADODARA_Seismic 2D Data Acquisition_Since inception_1782196863552.xlsx"

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql+asyncpg://ongc_user:ongc_pass@localhost:5433/ongc_db"
)

GP_RE = re.compile(r'GP[-]?(\d+[A-Z]?)', re.IGNORECASE)

def extract_gp(agency):
    if not agency:
        return None
    agency = str(agency).strip()
    m = GP_RE.search(agency)
    if m:
        code = m.group(0).upper()
        code = code.replace("GP0-", "GP-").replace("GP0", "GP-")
        if code.startswith("GP") and not code.startswith("GP-"):
            code = "GP-" + code[2:]
        return code
    return None

async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    engine = create_async_engine(DATABASE_URL)
    async with engine.begin() as conn:
        pass 

    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    inserted = 0
    skipped = 0
    new_gp_codes = set()
    new_blocks = set()
    new_categories = set()
    new_seasons = set()
    sheets_data = [
        ("Cambay", wb["Cambay"],
         {"season": 2, "sig": 3, "survey_type": 4, "project_name": 5,
          "agency": 6, "onland": 7, "volume": 8, "remarks": 9}),
        ("Rajasthan", wb["Rajasthan"],
         {"season": 2, "sig": 3, "survey_type": 4, "project_name": 5,
          "agency": 6, "onland": 7, "volume": 8, "remarks": 9}),
    ]

    async with session_factory() as session:
        for sheet_name, ws, cols in sheets_data:
            print(f"\nProcessing sheet: {sheet_name} ({ws.max_row - 1} total rows)")
            for r in range(2, ws.max_row + 1):
                season_val = ws.cell(row=r, column=cols["season"]).value
                if season_val is None:
                    skipped += 1
                    continue

                season = str(season_val).strip()
                sig = ws.cell(row=r, column=cols["sig"]).value
                survey_type = ws.cell(row=r, column=cols["survey_type"]).value
                project_name = ws.cell(row=r, column=cols["project_name"]).value
                agency = ws.cell(row=r, column=cols["agency"]).value
                onland = ws.cell(row=r, column=cols["onland"]).value
                volume = ws.cell(row=r, column=cols["volume"]).value
                remarks = ws.cell(row=r, column=cols["remarks"]).value

                if not project_name:
                    skipped += 1
                    continue

                project_name = str(project_name).strip()
                gp_code = extract_gp(agency)
                agency_str = str(agency).strip() if agency else None
                survey_type_str = str(survey_type).strip() if survey_type else "2D"
                sig_str = str(sig).strip() if sig else None
                volume_str = str(volume).strip() if volume is not None else None
                remarks_str = str(remarks).strip() if remarks else None
                onland_str = str(onland).strip() if onland else None

                new_seasons.add(season)

                project = Project(
                    project_name=project_name,
                    number=sig_str,
                    survey_type=survey_type_str,
                    contractor_name=agency_str,
                    area_name=sheet_name,
                    gp_code=gp_code,
                    year_field_season=season,
                    target_vs_achievement=volume_str,
                    project_highlights=remarks_str,
                    location=onland_str,
                    status="Historical",
                )
                session.add(project)
                inserted += 1

                if gp_code and gp_code not in ["GP-03","GP-06","GP-15","GP-16","GP-36","GP-61","GP-81"]:
                    new_gp_codes.add(gp_code)

            new_blocks.add(sheet_name)

        await session.commit()
        print(f"\nInserted: {inserted}, Skipped: {skipped}")

        existing_types = set()
        result = await session.execute(select(Lookup.type, Lookup.value).where(Lookup.is_active == True))
        for row in result:
            existing_types.add((row.type, row.value))

        lookup_inserts = []
        for gp in sorted(new_gp_codes):
            if ("section", gp) not in existing_types:
                lookup_inserts.append(Lookup(type="section", value=gp, sort_order=0))
        for b in sorted(new_blocks):
            if ("block", b) not in existing_types:
                lookup_inserts.append(Lookup(type="block", value=b, sort_order=0))
        if ("category", "Historical Seismic Data") not in existing_types:
            lookup_inserts.append(Lookup(type="category", value="Historical Seismic Data", sort_order=0))

        for lu in lookup_inserts:
            session.add(lu)
        if lookup_inserts:
            await session.commit()
            print(f"Added {len(lookup_inserts)} new lookup entries: {[lu.value for lu in lookup_inserts]}")

    await engine.dispose()
    wb.close()

if __name__ == "__main__":
    asyncio.run(main())
