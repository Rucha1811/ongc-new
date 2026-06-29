from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.database import get_db
from app.models.base import AcquisitionTarget, ManpowerEmployee, User
from app.auth.deps import get_current_user
from typing import Optional
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

MONTH_COLS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]
MONTH_COLS_ACH = ["apr_ach", "may_ach", "jun_ach", "jul_ach", "aug_ach", "sep_ach", "oct_ach", "nov_ach", "dec_ach", "jan_ach", "feb_ach", "mar_ach"]
MONTH_LABELS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


@router.get("/acquisition-targets")
async def list_targets(
    financial_year: Optional[str] = None,
    type: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget).order_by(AcquisitionTarget.project_name)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if type:
        q = q.where(AcquisitionTarget.type == type.upper())
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    return r.scalars().all()


@router.get("/acquisition-targets/export")
async def export_targets(
    project_name: Optional[str] = None,
    financial_year: Optional[str] = None,
    type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget).order_by(AcquisitionTarget.project_name, AcquisitionTarget.type, AcquisitionTarget.financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if type:
        q = q.where(AcquisitionTarget.type == type.upper())
    r = await db.execute(q)
    rows = r.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Acquisition Targets"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Project", "Project Type", "BE/RE", "FY", "Basin",
               "Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
               "Total", "Achieved", "%"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    for i, row in enumerate(rows, 2):
        vals = [
            row.project_name, row.project_type or "", row.type or "",
            row.financial_year or "", row.basin or "",
        ]
        total = 0
        for m in MONTH_COLS:
            v = getattr(row, m) or 0
            vals.append(v)
            total += v
        achieved = sum(getattr(row, m+"_ach") or 0 for m in MONTH_COLS)
        vals.append(total)
        vals.append(achieved)
        vals.append(round((achieved / total * 100) if total > 0 else 0, 1))

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = thin_border
            if col >= 6:
                c.alignment = Alignment(horizontal="right")

    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["E"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"acquisition_targets{'_'+project_name if project_name else ''}{'_'+financial_year if financial_year else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/acquisition-targets/{target_id}")
async def get_target(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    return t


@router.post("/acquisition-targets")
async def create_target(data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = AcquisitionTarget(**{k: v for k, v in data.items() if hasattr(AcquisitionTarget, k)})
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return t


@router.put("/acquisition-targets/{target_id}")
async def update_target(target_id: int, data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    for k, v in data.items():
        if hasattr(t, k):
            setattr(t, k, v)
    await db.commit()
    await db.refresh(t)
    return t


@router.delete("/acquisition-targets/{target_id}", status_code=204)
async def delete_target(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    await db.delete(t)
    await db.commit()


@router.get("/acquisition-targets/analytics/monthly")
async def monthly_analytics(
    financial_year: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    rows = r.scalars().all()

    result = {}
    for label in MONTH_LABELS:
        result[label] = {"be_target": 0, "be_achieved": 0, "re_target": 0, "re_achieved": 0}

    for row in rows:
        is_be = row.type == "BE"
        for i, month in enumerate(MONTH_COLS):
            label = MONTH_LABELS[i]
            tgt = getattr(row, month) or 0
            ach = getattr(row, MONTH_COLS_ACH[i]) or 0
            if is_be:
                result[label]["be_target"] += tgt
                result[label]["be_achieved"] += ach
            else:
                result[label]["re_target"] += tgt
                result[label]["re_achieved"] += ach

    return result


@router.get("/acquisition-targets/analytics/yearly")
async def yearly_analytics(
    financial_year: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    rows = r.scalars().all()

    be_target = 0
    be_achieved = 0
    re_target = 0
    re_achieved = 0

    for row in rows:
        if row.type == "BE":
            be_target += row.total or 0
            be_achieved += row.total_ach or 0
        else:
            re_target += row.total or 0
            re_achieved += row.total_ach or 0

    return {
        "be": {"target": be_target, "achieved": be_achieved},
        "re": {"target": re_target, "achieved": re_achieved},
    }


# ── Manpower Employees ──

@router.get("/manpower-employees")
async def list_manpower(
    section: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(ManpowerEmployee).order_by(ManpowerEmployee.section, ManpowerEmployee.sl_no)
    if section:
        q = q.where(ManpowerEmployee.section == section)
    r = await db.execute(q)
    return r.scalars().all()


@router.get("/manpower-employees/sections")
async def list_manpower_sections(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    r = await db.execute(select(ManpowerEmployee.section).distinct().order_by(ManpowerEmployee.section))
    return [row[0] for row in r]


@router.get("/manpower-employees/summary")
async def manpower_summary(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    r = await db.execute(select(ManpowerEmployee))
    rows = r.scalars().all()
    by_section = {}
    by_level = {}
    for m in rows:
        sec = m.section or "Unknown"
        by_section[sec] = by_section.get(sec, 0) + 1
        lvl = m.level or "Unknown"
        by_level[lvl] = by_level.get(lvl, 0) + 1
    return {"total": len(rows), "by_section": by_section, "by_level": by_level}


@router.post("/manpower-employees")
async def create_manpower(data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    m = ManpowerEmployee(**{k: v for k, v in data.items() if hasattr(ManpowerEmployee, k)})
    db.add(m)
    await db.commit()
    await db.refresh(m)
    return m


@router.delete("/manpower-employees/{emp_id}", status_code=204)
async def delete_manpower(emp_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    m = await db.get(ManpowerEmployee, emp_id)
    if not m:
        raise HTTPException(404, "Manpower record not found")
    await db.delete(m)
    await db.commit()
