from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.database import get_db
from app.models.base import Project, ProjectEvent, ProjectDocument, File as FileModel, User, Lookup
from app.auth.deps import get_current_user
from app.config import settings
import os, shutil, json, re
from datetime import date, datetime
from io import BytesIO
import openpyxl

router = APIRouter()

def _match_section(section: str, area_filter: str) -> bool:
    if not area_filter:
        return False
    parts = area_filter.split("/")
    first = parts[0].strip()
    prefix = ""
    if first.startswith("GP-"):
        prefix = "GP-"
    for p in parts:
        p = p.strip()
        if p == section:
            return True
        if prefix and not p.startswith(prefix) and prefix + p == section:
            return True
    return False

def parse_date(d):
    if not d:
        return None
    try:
        return date.fromisoformat(d)
    except:
        return None

def try_float(v):
    if not v:
        return None
    try:
        return float(v)
    except:
        return None

COLUMN_SYNONYMS = {
    "project_name": ["project name", "name of the area", "name of the area / project", "area/project name", "project", "area name"],
    "number": ["sig no", "sig. no.", "sig no.", "si no", "si no ", "si. no.", "number", "survey no"],
    "survey_type": ["type of survey", "survey type", "type"],
    "contractor_name": ["data acquired by", "data acquired by (agency name)", "agency", "agency(acquisition)", "contractor", "contractor name"],
    "area_name": ["area", "region", "block"],
    "section": ["section", "gp", "gp code"],
    "gp_code": ["gp", "gp code", "gpxx"],
    "year_field_season": ["field season", "season", "year"],
    "target_vs_achievement": ["volume", "volume  glk/skm", "volume glk/skm", "target", "achievement"],
    "project_highlights": ["remarks", "remarks (no of lines)", "highlights", "notes"],
    "location": ["onland/offshore", "location", "onland", "offshore", "area type"],
    "category": ["category", "classification"],
    "party_chief": ["party chief", "chief", "party"],
    "start_date": ["start date", "start", "commencement"],
    "end_date": ["end date", "end", "completion"],
    "status": ["status", "project status"],
}

def normalize_header(col):
    s = col.strip().lower()
    s = re.sub(r'\s+', ' ', s)
    for field, synonyms in COLUMN_SYNONYMS.items():
        for syn in synonyms:
            if s == syn or s.startswith(syn) or syn.startswith(s):
                return field
    return None

@router.get("/")
async def list_projects(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    role_name = user.role.name if user.role else "viewer"
    if role_name == "admin":
        result = await db.execute(select(Project).order_by(Project.created_at.desc()))
    elif role_name == "ops_manager":
        result = await db.execute(
            select(Project).order_by(Project.created_at.desc())
        )
    else:
        result = await db.execute(
            select(Project).order_by(Project.created_at.desc())
        )
    projects = result.scalars().all()

    if role_name not in ("admin", "ops_manager", "data_creator"):
        filtered = []
        for p in projects:
            match = False
            if p.section and user.area and _match_section(p.section, user.area):
                match = True
            elif p.section and user.section and p.section == user.section:
                match = True
            if match:
                filtered.append(p)
        projects = filtered
    out = []
    for p in projects:
        out.append({
            "id": p.id,
            "project_name": p.project_name,
            "number": p.number,
            "survey_type": p.survey_type,
            "contractor_name": p.contractor_name,
            "area_name": p.area_name,
            "section": p.section,
            "gp_code": p.gp_code,
            "category": p.category,
            "location": p.location,
            "party_chief": p.party_chief,
            "year_field_season": p.year_field_season,
            "status": p.status,
            "created_at": str(p.created_at) if p.created_at else None,
        })
    return out

@router.get("/{project_id}")
async def get_project(project_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(select(Project).where(Project.id == project_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Project not found")
    events_result = await db.execute(select(ProjectEvent).where(ProjectEvent.project_id == project_id).order_by(ProjectEvent.event_date))
    events = [{"event_date": str(e.event_date), "description": e.description} for e in events_result.scalars().all()]
    docs_result = await db.execute(select(ProjectDocument).where(ProjectDocument.project_id == project_id))
    docs = [{"id": d.id, "file_name": d.file_name, "file_type": d.file_type, "category": d.category} for d in docs_result.scalars().all()]
    return {
        "id": p.id,
        "project_name": p.project_name,
        "number": p.number,
        "survey_type": p.survey_type,
        "contractor_name": p.contractor_name,
        "area_name": p.area_name,
        "section": p.section,
        "gp_code": p.gp_code,
        "party_chief": p.party_chief,
        "year_field_season": p.year_field_season,
        "start_date": str(p.start_date) if p.start_date else None,
        "end_date": str(p.end_date) if p.end_date else None,
        "project_period": p.project_period,
        "target_vs_achievement": p.target_vs_achievement,
        "survey_objective": p.survey_objective,
        "xy_coordinates": p.xy_coordinates,
        "kml_file_path": p.kml_file_path,
        "survey_grid_params": p.survey_grid_params,
        "acquisition_geometry": p.acquisition_geometry,
        "instrument_parameters": p.instrument_parameters,
        "sensor_type": p.sensor_type,
        "source_parameters": p.source_parameters,
        "total_cost": p.total_cost,
        "per_unit_cost": p.per_unit_cost,
        "project_highlights": p.project_highlights,
        "category": p.category,
        "location": p.location,
        "project_map_path": p.project_map_path,
        "status": p.status,
        "created_at": str(p.created_at) if p.created_at else None,
        "updated_at": str(p.updated_at) if p.updated_at else None,
        "events": events,
        "documents": docs,
    }

@router.post("/create", status_code=201)
async def create_project(
    project_name: str = Form(...),
    number: str = Form(None),
    survey_type: str = Form(None),
    contractor_name: str = Form(None),
    area_name: str = Form(None),
    section: str = Form(None),
    gp_code: str = Form(None),
    party_chief: str = Form(None),
    year_field_season: str = Form(None),
    start_date: str = Form(None),
    end_date: str = Form(None),
    project_period: str = Form(None),
    target_vs_achievement: str = Form(None),
    survey_objective: str = Form(None),
    xy_coordinates: str = Form(None),
    survey_grid_params: str = Form(None),
    acquisition_geometry: str = Form(None),
    instrument_parameters: str = Form(None),
    sensor_type: str = Form(None),
    source_parameters: str = Form(None),
    total_cost: float = Form(None),
    per_unit_cost: float = Form(None),
    project_highlights: str = Form(None),
    category: str = Form(None),
    location: str = Form(None),
    events_json: str = Form("[]"),
    kml_file: UploadFile = File(None),
    project_map: UploadFile = File(None),
    related_files: str = Form("[]"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager"):
        raise HTTPException(status_code=403, detail="Only admins and ops managers can create projects")
    upload_dir = getattr(settings, "PROJECT_UPLOAD_DIR", os.path.join(settings.UPLOAD_DIR, "projects"))
    os.makedirs(upload_dir, exist_ok=True)

    kml_path = None
    if kml_file and kml_file.filename:
        kml_path = os.path.join(upload_dir, f"kml_{project_name}_{kml_file.filename}")
        with open(kml_path, "wb") as f:
            shutil.copyfileobj(kml_file.file, f)

    map_path = None
    if project_map and project_map.filename:
        map_path = os.path.join(upload_dir, f"map_{project_name}_{project_map.filename}")
        with open(map_path, "wb") as f:
            shutil.copyfileobj(project_map.file, f)

    def parse_date(d: str):
        if not d:
            return None
        try:
            return date.fromisoformat(d)
        except:
            return None

    project = Project(
        project_name=project_name,
        number=number,
        survey_type=survey_type,
        contractor_name=contractor_name,
        area_name=area_name,
        section=section,
        gp_code=gp_code,
        party_chief=party_chief,
        year_field_season=year_field_season,
        start_date=parse_date(start_date),
        end_date=parse_date(end_date),
        project_period=project_period,
        target_vs_achievement=target_vs_achievement,
        survey_objective=survey_objective,
        xy_coordinates=xy_coordinates,
        kml_file_path=kml_path,
        survey_grid_params=survey_grid_params,
        acquisition_geometry=acquisition_geometry,
        instrument_parameters=instrument_parameters,
        sensor_type=sensor_type,
        source_parameters=source_parameters,
        total_cost=total_cost,
        per_unit_cost=per_unit_cost,
        project_highlights=project_highlights,
        category=category,
        location=location,
        project_map_path=map_path,
        created_by=user.id,
    )
    db.add(project)
    await db.flush()

    try:
        events_data = json.loads(events_json)
        for ev in events_data:
            ed = ev.get("event_date")
            desc = ev.get("description")
            if ed and desc:
                pe = ProjectEvent(project_id=project.id, event_date=parse_date(ed), description=desc)
                db.add(pe)
    except:
        pass

    try:
        related = json.loads(related_files)
        for rf in related:
            if isinstance(rf, dict):
                pd = ProjectDocument(
                    project_id=project.id,
                    file_name=rf.get("file_name", "unknown"),
                    file_path=rf.get("file_path"),
                    file_type=rf.get("file_type"),
                    category="related_doc",
                )
                db.add(pd)
    except:
        pass

    await db.commit()
    return {"id": project.id, "project_name": project.project_name, "msg": "Project created"}

@router.patch("/{project_id}")
async def update_project(project_id: int, data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    role_name = user.role.name if user.role else ""
    if role_name not in ("admin", "ops_manager"):
        raise HTTPException(status_code=403, detail="Only admin/ops_manager can update project")
    result = await db.execute(select(Project).where(Project.id == project_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Project not found")
    for key, val in data.items():
        if hasattr(p, key) and key not in ("id", "created_at", "created_by"):
            setattr(p, key, val)
    await db.commit()
    return {"msg": "Project updated"}

@router.post("/{project_id}/upload")
async def upload_project_file(
    project_id: int,
    file: UploadFile = File(...),
    file_type: str = Form(...),
    data_type: str = Form(None),
    category: str = Form(None),
    classification: str = Form(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager", "data_creator"):
        raise HTTPException(status_code=403, detail="You do not have permission to upload files")
    result = await db.execute(select(Project).where(Project.id == project_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Project not found")

    allowed_ext = {"pdf","docx","xlsx","ppt","pptx","txt","dat","csv","zip"}
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"File type '.{ext}' not allowed")

    start_year = str(p.start_date.year) if p.start_date else (p.year_field_season or "")
    section_part = p.section or ""
    proj_number = p.number or ""
    dtype = data_type or ""
    area = p.area_name or ""
    contractor = p.contractor_name or ""

    doc_count = await db.execute(select(func.count(ProjectDocument.id)).where(ProjectDocument.project_id == project_id))
    seq = (doc_count.scalar() or 0) + 1

    parts = [start_year, section_part, proj_number, dtype, area, contractor]
    clean_name = "_".join(p for p in parts if p).replace(" ", "_") + f"_{seq:03d}.{ext}"

    section_safe = p.section.replace("/", "_").replace(" ", "_") if p.section else "General"
    project_safe = p.project_name.replace("/", "_").replace(" ", "_") if p.project_name else "Uncategorized"
    classification_safe = classification.replace("/", "_").replace(" ", "_") if classification else "Unclassified"
    project_upload_dir = os.path.join(settings.UPLOAD_DIR, project_safe, section_safe, classification_safe)
    os.makedirs(project_upload_dir, exist_ok=True)
    file_path = os.path.join(project_upload_dir, clean_name)
    contents = await file.read()
    with open(file_path, "wb") as fh:
        fh.write(contents)

    pd = ProjectDocument(
        project_id=project_id,
        file_name=clean_name,
        file_path=file_path,
        file_type=file_type.upper(),
        category=category,
    )
    db.add(pd)

    from app.activity_utils import log_activity
    auto_approved = user.role.name == "admin" if user.role else False
    db_file = FileModel(
        file_name=clean_name,
        file_type=file_type.upper(),
        project_name=p.project_name,
        sig_number=p.number,
        data_type=data_type,
        section=p.section,
        category=category,
        classification=classification,
        contractor_name=p.contractor_name,
        status="Approved" if auto_approved else "Pending",
        uploaded_by=user.id,
        upload_date=datetime.utcnow(),
        file_size=f"{len(contents)/1024/1024:.2f} MB",
        file_path=file_path,
    )
    db.add(db_file)
    await db.commit()

    if auto_approved:
        from app.models.base import Approval
        approval = Approval(file_id=db_file.id, action="approved", action_by=user.id, action_at=datetime.utcnow(), comment="Auto-approved (admin upload)")
        db.add(approval)
        await db.commit()

    await log_activity(db, user.id, "upload", "file", db_file.id, f"Uploaded to project '{p.project_name}': {clean_name} ({classification})")
    await db.commit()
    return {"msg": "File uploaded to project", "file_name": clean_name}

@router.delete("/{project_id}")
async def delete_project(project_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    role_name = user.role.name if user.role else "viewer"
    if role_name != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete projects")
    result = await db.execute(select(Project).where(Project.id == project_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Project not found")
    await db.delete(p)
    await db.commit()
    return {"msg": "Project deleted"}

@router.post("/upload-excel/preview")
async def excel_preview(
    file: UploadFile = File(...),
    sheet_name: str = Form(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager"):
        raise HTTPException(403, "Only admin/ops_manager can upload Excel")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported")
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    sheets = wb.sheetnames
    if sheet_name and sheet_name in sheets:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title
    headers = [c.value for c in ws[1] if c.value]
    auto_mapping = {}
    unmapped = []
    for h in headers:
        field = normalize_header(h)
        if field:
            auto_mapping[h] = field
        else:
            unmapped.append(h)
    preview = []
    for r in range(2, min(7, ws.max_row + 1)):
        row_data = {}
        for ci, c in enumerate(ws[r]):
            if ci < len(headers):
                row_data[headers[ci]] = str(c.value) if c.value is not None else ""
        if any(row_data.values()):
            preview.append(row_data)
    project_name_header = next((k for k,v in auto_mapping.items() if v=="project_name"), None)
    existing = set()
    if project_name_header and project_name_header in headers:
        result = await db.execute(select(Project.project_name))
        existing = set(row[0] for row in result if row[0])
    dup_rows = 0
    if project_name_header and project_name_header in headers:
        col_idx = headers.index(project_name_header)
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx + 1).value
            if val and str(val).strip() in existing:
                dup_rows += 1
    wb.close()
    return {
        "auto": len(unmapped) == 0,
        "sheet_name": sheet_name,
        "sheets": sheets,
        "columns": headers,
        "row_count": max(0, ws.max_row - 1),
        "preview": preview,
        "auto_mapping": auto_mapping,
        "unmapped": unmapped,
        "duplicate_count": dup_rows,
    }

@router.post("/upload-excel/import", status_code=201)
async def excel_import(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    sheet_name: str = Form(None),
    conflict: str = Form("skip"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager"):
        raise HTTPException(403, "Only admin/ops_manager can import Excel")
    mapping_dict = json.loads(mapping)
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    headers = [c.value for c in ws[1] if c.value]
    col_idx = {h: i for i, h in enumerate(headers)}

    valid_fields = {
        "project_name","number","survey_type","contractor_name","area_name",
        "section","gp_code","party_chief","year_field_season","start_date","end_date",
        "project_period","target_vs_achievement","survey_objective","xy_coordinates",
        "survey_grid_params","acquisition_geometry","instrument_parameters","sensor_type",
        "source_parameters","total_cost","per_unit_cost","project_highlights",
        "category","location","status",
    }

    project_name_field = next((k for k,v in mapping_dict.items() if v=="project_name"), None)
    existing_names = set()
    if conflict == "skip" and project_name_field:
        result = await db.execute(select(Project.project_name))
        existing_names = set(row[0] for row in result if row[0])

    imported = 0
    skipped = 0
    for r in range(2, ws.max_row + 1):
        row_data = {}
        excel_col = None
        has_data = False
        for col_name, field_name in mapping_dict.items():
            if col_name in col_idx and field_name in valid_fields:
                val = ws.cell(row=r, column=col_idx[col_name] + 1).value
                if val is not None:
                    row_data[field_name] = str(val).strip()
                    if field_name == "project_name":
                        has_data = True
        if not has_data:
            continue

        pname = row_data.get("project_name", f"Imported-{r}")
        if pname in existing_names:
            skipped += 1
            continue
        project = Project(
            project_name=row_data.get("project_name", f"Imported-{r}"),
            number=row_data.get("number"),
            survey_type=row_data.get("survey_type"),
            contractor_name=row_data.get("contractor_name"),
            area_name=row_data.get("area_name"),
            section=row_data.get("section"),
            gp_code=row_data.get("gp_code"),
            party_chief=row_data.get("party_chief"),
            year_field_season=row_data.get("year_field_season"),
            start_date=parse_date(row_data.get("start_date")),
            end_date=parse_date(row_data.get("end_date")),
            project_period=row_data.get("project_period"),
            target_vs_achievement=row_data.get("target_vs_achievement"),
            survey_objective=row_data.get("survey_objective"),
            xy_coordinates=row_data.get("xy_coordinates"),
            survey_grid_params=row_data.get("survey_grid_params"),
            acquisition_geometry=row_data.get("acquisition_geometry"),
            instrument_parameters=row_data.get("instrument_parameters"),
            sensor_type=row_data.get("sensor_type"),
            source_parameters=row_data.get("source_parameters"),
            total_cost=try_float(row_data.get("total_cost")),
            per_unit_cost=try_float(row_data.get("per_unit_cost")),
            project_highlights=row_data.get("project_highlights"),
            category=row_data.get("category"),
            location=row_data.get("location"),
            status=row_data.get("status", "Historical"),
            created_by=user.id,
        )
        db.add(project)
        imported += 1
    await db.commit()
    wb.close()
    return {"imported": imported, "skipped": skipped, "msg": f"{imported} projects imported, {skipped} duplicates skipped"}
