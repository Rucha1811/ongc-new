import asyncio, os
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from app.models.base import ManpowerEmployee
from sqlalchemy import text
import openpyxl

EXCEL_PATH = "/Users/ruchatejaskumargandhi/Desktop/ONGC 3/Manpower_Geophysical_Services-2026.06.22_035425.xlsx"
DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql+asyncpg://ongc_user:ongc_pass@localhost:5432/ongc_db")

SECTION_KEYWORDS = ["Base Office", "GP-", "RCC", "REL"]

async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Manpower Status GPS-2026"]

    engine = create_async_engine(DATABASE_URL)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with session_factory() as session:
        await session.execute(text("DELETE FROM manpower_employees"))
        await session.commit()

        current_section = None
        inserted = 0

        for r in range(3, ws.max_row + 1):
            col1 = ws.cell(row=r, column=1).value
            col2 = ws.cell(row=r, column=2).value
            col3 = ws.cell(row=r, column=3).value

            col1_str = str(col1).strip().replace("\xa0", "") if col1 else ""

            # Detect section header
            if col1:
                is_section = any(col1_str.startswith(k) for k in SECTION_KEYWORDS)
                if is_section:
                    current_section = col1_str
                    continue

            if not col3:
                continue

            # Skip column header row
            if col1_str.lower().startswith("sl"):
                continue

            # Must have a numeric sl no
            try:
                sl_no = int(float(col1_str))
            except (ValueError, TypeError):
                continue

            name = str(col3).strip().replace("\xa0", "")
            desig = str(ws.cell(row=r, column=4).value or "").strip()
            mobile = str(ws.cell(row=r, column=5).value or "").strip()
            level = str(ws.cell(row=r, column=6).value or "").strip()
            crc = str(ws.cell(row=r, column=7).value or "").strip()
            assignment = str(ws.cell(row=r, column=8).value or "").strip()
            cpf_no = str(int(float(col2))) if col2 else None

            emp = ManpowerEmployee(
                section=current_section or "Unknown",
                basin="Western",
                sl_no=sl_no,
                cpf_no=cpf_no,
                name=name,
                designation=desig if desig else None,
                mobile=mobile if mobile else None,
                level=level if level else None,
                crc=crc if crc else None,
                assignment=assignment if assignment else None,
            )
            session.add(emp)
            inserted += 1

        await session.commit()
        print(f"Imported {inserted} manpower records")
        result = await session.execute(text("SELECT section, count(*) FROM manpower_employees GROUP BY section ORDER BY section"))
        for row in result:
            print(f"  {row[0]}: {row[1]}")

    await engine.dispose()
    wb.close()

if __name__ == "__main__":
    asyncio.run(main())
