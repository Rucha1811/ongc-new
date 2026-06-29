#!/usr/bin/env python3
"""Import BE/RE acquisition targets from the Excel sheet."""

import asyncio, os, sys
sys.path.insert(0, os.path.dirname(__file__))

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import select
from app.models.base import AcquisitionTarget
import openpyxl
from collections import defaultdict

import glob
_candidates = glob.glob(os.path.expanduser("~/Desktop/ONGC 3/*Acquisition*Targets*032908*"))
EXCEL_PATH = _candidates[0] if _candidates else os.path.join(
    os.path.dirname(__file__), "..",
    "2026.04.20_Monthly Acquisition Targets for  RE2026-27_BE2027-28_032908.xlsx")

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql+asyncpg://ongc_user:ongc_pass@localhost:5432/ongc_db")

MONTHS = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]

def parse_sheet(ws):
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        sl = row[0]
        area = row[1]
        annual_target = row[2]
        if not isinstance(annual_target, (int, float)):
            continue
        if sl is None and area is None:
            continue
        if sl is not None and area:
            rows.append(row)
    return rows

def sheet_type_fy(sheet_name):
    parts = sheet_name.strip().split()
    return parts[0], parts[1]

async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    engine = create_async_engine(DATABASE_URL)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with session_factory() as session:
        count = 0
        for sheet_name in wb.sheetnames:
            be_re, fy = sheet_type_fy(sheet_name)
            ws = wb[sheet_name]
            data_rows = parse_sheet(ws)

            aggregated = defaultdict(lambda: {"months": [0]*12, "total": 0})

            for row in data_rows:
                area = str(row[1]).strip()
                annual_target = row[2] if row[2] is not None else 0
                row_total = row[15] if row[15] is not None else 0

                    area = area.replace("\n", " ").replace("\r", " ").strip()
                    monthly = [0]*12
                for i in range(12):
                    v = row[3+i]
                    monthly[i] = float(v) if v is not None else 0

                key = (area, be_re, fy)
                agg = aggregated[key]
                for i in range(12):
                    agg["months"][i] += monthly[i]
                agg["total"] += float(row_total or annual_target or 0)

            for (area, be_re, fy), agg in sorted(aggregated.items()):
                result = await session.execute(
                    select(AcquisitionTarget).where(
                        AcquisitionTarget.project_name == area,
                        AcquisitionTarget.type == be_re,
                        AcquisitionTarget.financial_year == fy,
                    )
                )
                existing = result.scalar_one_or_none()

                if existing:
                    print(f"  Updating: {area} / {be_re} / {fy}")
                    existing.total = agg["total"]
                    for i, m in enumerate(MONTHS):
                        setattr(existing, m, agg["months"][i])
                else:
                    print(f"  Creating: {area} / {be_re} / {fy}")
                    rec = AcquisitionTarget(
                        project_name=area,
                        project_type=None,
                        type=be_re,
                        financial_year=fy,
                        basin="WON",
                        total=agg["total"],
                    )
                    for i, m in enumerate(MONTHS):
                        setattr(rec, m, agg["months"][i])
                    session.add(rec)
                count += 1

        await session.commit()
        print(f"\nDone! {count} records imported.")

    await engine.dispose()

if __name__ == "__main__":
    asyncio.run(main())
